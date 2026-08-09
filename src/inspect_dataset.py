from pathlib import Path

from openpyxl import load_workbook

from src.contracts.paths import RAW_DATASET

DATASET_PATH = RAW_DATASET


def inspect_workbook(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"Dataset not found: {path}")

    print("=" * 80)
    print("MAYASARI BAKERY DATASET INSPECTION")
    print("=" * 80)
    print(f"File : {path}")
    print(f"Size : {path.stat().st_size / 1024 / 1024:.2f} MB")
    print()

    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )

    for worksheet in workbook.worksheets:
        print("=" * 80)
        print(f"SHEET: {worksheet.title}")
        print("=" * 80)

        rows = worksheet.iter_rows(values_only=True)

        try:
            headers = next(rows)
        except StopIteration:
            print("Empty sheet.")
            continue

        print()
        print("COLUMNS")
        print("-" * 80)

        for index, header in enumerate(headers, start=1):
            print(f"{index:>2}. {header}")

        print()
        print("SAMPLE DATA")
        print("-" * 80)

    for sample_count, row in enumerate(rows, start=1):
        print(row)

        if sample_count >= 3:
            break

        print()

    workbook.close()

    print("=" * 80)
    print("Inspection completed successfully.")
    print("=" * 80)


if __name__ == "__main__":
    inspect_workbook(DATASET_PATH)
