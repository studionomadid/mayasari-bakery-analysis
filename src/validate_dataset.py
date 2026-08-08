from pathlib import Path

from openpyxl import load_workbook


DATASET_PATH = Path(
    "data/raw/mayasari_bakery_2025_synthetic.xlsx"
)


def load_sheet(workbook, sheet_name):
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)

    headers = list(next(rows))
    records = [dict(zip(headers, row)) for row in rows]

    return records


def check_duplicates(records, key):
    values = [
        row[key]
        for row in records
        if row[key] is not None
    ]

    return len(values) - len(set(values))


def check_sales_line_duplicates(records):
    keys = [
        (
            row["transaction_id"],
            row["line_id"],
        )
        for row in records
    ]

    return len(keys) - len(set(keys))


def main():
    if not DATASET_PATH.exists():
        raise FileNotFoundError(
            f"Dataset not found: {DATASET_PATH}"
        )

    workbook = load_workbook(
        DATASET_PATH,
        read_only=True,
        data_only=True,
    )

    customers = load_sheet(workbook, "customers")
    products = load_sheet(workbook, "products")
    sales = load_sheet(workbook, "sales")
    expenses = load_sheet(workbook, "expenses")
    monthly_kpi = load_sheet(workbook, "monthly_kpi")

    print("=" * 80)
    print("MAYASARI BAKERY DATA VALIDATION")
    print("=" * 80)

    # ------------------------------------------------------------------
    # Record counts
    # ------------------------------------------------------------------

    print("\n[1] RECORD COUNTS")
    print("-" * 80)

    print(f"customers  : {len(customers):,}")
    print(f"products   : {len(products):,}")
    print(f"sales      : {len(sales):,}")
    print(f"expenses   : {len(expenses):,}")
    print(f"monthly_kpi: {len(monthly_kpi):,}")

    # ------------------------------------------------------------------
    # Primary key validation
    # ------------------------------------------------------------------

    print("\n[2] PRIMARY KEY VALIDATION")
    print("-" * 80)

    customer_duplicates = check_duplicates(
        customers,
        "customer_id",
    )

    product_duplicates = check_duplicates(
        products,
        "product_id",
    )

    sales_line_duplicates = check_sales_line_duplicates(
        sales
    )

    expense_duplicates = check_duplicates(
        expenses,
        "expense_id",
    )

    print(
        f"customer_id duplicates              : "
        f"{customer_duplicates}"
    )

    print(
        f"product_id duplicates               : "
        f"{product_duplicates}"
    )

    print(
        f"(transaction_id, line_id) duplicates: "
        f"{sales_line_duplicates}"
    )

    print(
        f"expense_id duplicates               : "
        f"{expense_duplicates}"
    )

    # ------------------------------------------------------------------
    # Transaction ID analysis
    # ------------------------------------------------------------------

    print("\n[3] TRANSACTION STRUCTURE")
    print("-" * 80)

    transaction_ids = [
        row["transaction_id"]
        for row in sales
        if row["transaction_id"] is not None
    ]

    unique_transactions = len(set(transaction_ids))

    print(
        f"Sales line records : {len(sales):,}"
    )

    print(
        f"Unique transactions: {unique_transactions:,}"
    )

    print(
        f"Average lines/transaction: "
        f"{len(sales) / unique_transactions:.2f}"
    )

    # ------------------------------------------------------------------
    # Foreign key validation
    # ------------------------------------------------------------------

    print("\n[4] FOREIGN KEY VALIDATION")
    print("-" * 80)

    customer_ids = {
        row["customer_id"]
        for row in customers
    }

    product_ids = {
        row["product_id"]
        for row in products
    }

    missing_customers = {
        row["customer_id"]
        for row in sales
        if row["customer_id"] not in customer_ids
    }

    missing_products = {
        row["product_id"]
        for row in sales
        if row["product_id"] not in product_ids
    }

    print(
        f"Sales with unknown customer_id : "
        f"{len(missing_customers)}"
    )

    print(
        f"Sales with unknown product_id  : "
        f"{len(missing_products)}"
    )

    # ------------------------------------------------------------------
    # Missing values
    # ------------------------------------------------------------------

    print("\n[5] MISSING VALUES")
    print("-" * 80)

    sheets = {
        "customers": customers,
        "products": products,
        "sales": sales,
        "expenses": expenses,
        "monthly_kpi": monthly_kpi,
    }

    for sheet_name, records in sheets.items():
        missing = {}

        if not records:
            continue

        for column in records[0]:
            count = sum(
                1
                for row in records
                if row[column] is None
            )

            if count:
                missing[column] = count

        if missing:
            print(f"{sheet_name}:")

            for column, count in missing.items():
                print(
                    f"  {column}: {count}"
                )

        else:
            print(
                f"{sheet_name}: no missing values"
            )

    # ------------------------------------------------------------------
    # Sales business rules
    # ------------------------------------------------------------------

    print("\n[6] SALES BUSINESS RULES")
    print("-" * 80)

    invalid_quantity = 0
    invalid_unit_price = 0
    invalid_discount_rate = 0
    invalid_discount_amount = 0
    invalid_net_sales = 0
    invalid_gross_sales = 0
    invalid_product_cost = 0
    invalid_gross_profit = 0

    for row in sales:
        quantity = row["quantity"]
        unit_price = row["unit_price"]
        discount_rate = row["discount_rate"]
        discount_amount = row["discount_amount"]
        net_sales = row["net_sales"]
        gross_sales = row["gross_sales"]
        product_cost = row["product_cost"]
        gross_profit = row["gross_profit"]

        if quantity is None or quantity <= 0:
            invalid_quantity += 1

        if unit_price is None or unit_price <= 0:
            invalid_unit_price += 1

        if (
            discount_rate is None
            or discount_rate < 0
            or discount_rate > 1
        ):
            invalid_discount_rate += 1

        if (
            quantity is None
            or unit_price is None
            or discount_rate is None
        ):
            continue

        expected_gross_sales = (
            quantity * unit_price
        )

        expected_discount_amount = (
            expected_gross_sales * discount_rate
        )

        expected_net_sales = (
            expected_gross_sales
            - expected_discount_amount
        )

        if product_cost is not None:
            expected_gross_profit = (
                expected_net_sales
                - product_cost
            )
        else:
            expected_gross_profit = None

        tolerance = 0.01

        if (
            discount_amount is None
            or abs(
                discount_amount
                - expected_discount_amount
            ) > tolerance
        ):
            invalid_discount_amount += 1

        if (
            gross_sales is None
            or abs(
                gross_sales
                - expected_gross_sales
            ) > tolerance
        ):
            invalid_gross_sales += 1

        if (
            net_sales is None
            or abs(
                net_sales
                - expected_net_sales
            ) > tolerance
        ):
            invalid_net_sales += 1

        if (
            product_cost is None
            or product_cost < 0
        ):
            invalid_product_cost += 1

        if (
            expected_gross_profit is None
            or gross_profit is None
            or abs(
                gross_profit
                - expected_gross_profit
            ) > tolerance
        ):
            invalid_gross_profit += 1

    print(
        f"Invalid quantity       : "
        f"{invalid_quantity}"
    )

    print(
        f"Invalid unit price     : "
        f"{invalid_unit_price}"
    )

    print(
        f"Invalid discount rate  : "
        f"{invalid_discount_rate}"
    )

    print(
        f"Invalid discount amount: "
        f"{invalid_discount_amount}"
    )

    print(
        f"Invalid gross sales    : "
        f"{invalid_gross_sales}"
    )

    print(
        f"Invalid net sales      : "
        f"{invalid_net_sales}"
    )

    print(
        f"Invalid product cost   : "
        f"{invalid_product_cost}"
    )

    print(
        f"Invalid gross profit   : "
        f"{invalid_gross_profit}"
    )

    # ------------------------------------------------------------------
    # Product master validation
    # ------------------------------------------------------------------

    print("\n[7] PRODUCT MASTER VALIDATION")
    print("-" * 80)

    invalid_product_price = 0
    invalid_product_cost = 0

    for row in products:
        if (
            row["price"] is None
            or row["price"] <= 0
        ):
            invalid_product_price += 1

        if (
            row["cost"] is None
            or row["cost"] < 0
        ):
            invalid_product_cost += 1

    print(
        f"Invalid product price : "
        f"{invalid_product_price}"
    )

    print(
        f"Invalid product cost  : "
        f"{invalid_product_cost}"
    )

    # ------------------------------------------------------------------
    # Expense validation
    # ------------------------------------------------------------------

    print("\n[8] EXPENSE VALIDATION")
    print("-" * 80)

    invalid_expenses = sum(
        1
        for row in expenses
        if (
            row["amount"] is None
            or row["amount"] < 0
        )
    )

    print(
        f"Invalid expense amount: "
        f"{invalid_expenses}"
    )

    # ------------------------------------------------------------------
    # Date validation
    # ------------------------------------------------------------------

    print("\n[9] DATE VALIDATION")
    print("-" * 80)

    invalid_sales_dates = sum(
        1
        for row in sales
        if row["transaction_date"] is None
    )

    invalid_expense_dates = sum(
        1
        for row in expenses
        if row["expense_date"] is None
    )

    invalid_customer_dates = sum(
        1
        for row in customers
        if row["registration_date"] is None
    )

    print(
        f"Missing sales dates    : "
        f"{invalid_sales_dates}"
    )

    print(
        f"Missing expense dates  : "
        f"{invalid_expense_dates}"
    )

    print(
        f"Missing customer dates : "
        f"{invalid_customer_dates}"
    )

    # ------------------------------------------------------------------
    # Monthly KPI validation
    # ------------------------------------------------------------------

    print("\n[10] MONTHLY KPI")
    print("-" * 80)

    months = [
        row["month"]
        for row in monthly_kpi
        if row["month"] is not None
    ]

    unique_months = len(set(months))

    print(
        f"KPI records  : "
        f"{len(monthly_kpi)}"
    )

    print(
        f"Unique months : "
        f"{unique_months}"
    )

    if (
        len(monthly_kpi) == 12
        and unique_months == 12
    ):
        print(
            "Monthly KPI coverage: PASS"
        )
    else:
        print(
            "Monthly KPI coverage: REVIEW"
        )

    workbook.close()

    # ------------------------------------------------------------------
    # Final validation status
    # ------------------------------------------------------------------

    all_checks_pass = all(
        [
            customer_duplicates == 0,
            product_duplicates == 0,
            sales_line_duplicates == 0,
            expense_duplicates == 0,
            len(missing_customers) == 0,
            len(missing_products) == 0,
            invalid_quantity == 0,
            invalid_unit_price == 0,
            invalid_discount_rate == 0,
            invalid_discount_amount == 0,
            invalid_gross_sales == 0,
            invalid_net_sales == 0,
            invalid_product_cost == 0,
            invalid_gross_profit == 0,
            invalid_product_price == 0,
            invalid_product_cost == 0,
            invalid_expenses == 0,
            invalid_sales_dates == 0,
            invalid_expense_dates == 0,
            invalid_customer_dates == 0,
            len(monthly_kpi) == 12,
            unique_months == 12,
        ]
    )

    print()
    print("=" * 80)

    if all_checks_pass:
        print("VALIDATION STATUS: PASS")
    else:
        print("VALIDATION STATUS: REVIEW")

    print("=" * 80)


if __name__ == "__main__":
    main()
